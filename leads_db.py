"""
Leads database: parse Excel base into SQLite for the sales funnel.
"""
import os
import sqlite3
import logging
from pathlib import Path
from dataclasses import dataclass

logger = logging.getLogger("leads_db")

DB_PATH = Path(__file__).parent / "leads.db"


@dataclass
class Lead:
    id: int | None
    company_name: str
    region: str
    city: str
    phone: str
    mobile: str
    email: str
    whatsapp: str
    telegram: str
    industry: str
    rating: float
    status: str = "new"  # new, called, interested, callback, refused, sent_wa, sent_email, done
    call_result: str = ""
    notes: str = ""


def init_db(db_path: str | Path | None = None) -> sqlite3.Connection:
    """Initialize database and create tables if needed."""
    path = str(db_path or DB_PATH)
    conn = sqlite3.connect(path, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("""
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT NOT NULL,
            region TEXT,
            city TEXT,
            phone TEXT,
            mobile TEXT,
            email TEXT,
            whatsapp TEXT,
            telegram TEXT,
            industry TEXT,
            rating REAL DEFAULT 0,
            status TEXT DEFAULT 'new',
            call_result TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS call_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lead_id INTEGER REFERENCES leads(id),
            call_type TEXT,
            result TEXT,
            duration_sec INTEGER,
            transcript TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS message_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lead_id INTEGER REFERENCES leads(id),
            channel TEXT,
            message_type TEXT,
            content TEXT,
            status TEXT,
            external_id TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_status ON leads(status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_industry ON leads(industry)")
    conn.commit()
    return conn


def import_excel(excel_path: str, db_path: str | None = None) -> int:
    """
    Import leads from Excel file into SQLite.
    Returns number of imported leads.
    """
    import openpyxl

    conn = init_db(db_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    total = 0

    # Skip summary sheet
    sheets = [s for s in wb.sheetnames if "сводн" not in s.lower()]

    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First row = headers
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]

        # Map column indices
        col_map = {}
        for i, h in enumerate(headers):
            if "название" in h:
                col_map["company_name"] = i
            elif h == "регион":
                col_map["region"] = i
            elif h == "город":
                col_map["city"] = i
            elif "телефон" in h and "мобильн" not in h:
                col_map["phone"] = i
            elif "мобильн" in h:
                col_map["mobile"] = i
            elif "email" in h:
                col_map["email"] = i
            elif "whatsapp" in h:
                col_map["whatsapp"] = i
            elif "telegram" in h:
                col_map["telegram"] = i
            elif "рейтинг" in h:
                col_map["rating"] = i

        for row in rows[1:]:
            if not row or not row[col_map.get("company_name", 0)]:
                continue

            def get(key):
                idx = col_map.get(key)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val else ""
                return ""

            company_name = get("company_name")
            if not company_name:
                continue

            # Clean phone numbers
            mobile = get("mobile")
            phone = get("phone")
            whatsapp = get("whatsapp")

            # WhatsApp field is often masked (****), use mobile as fallback
            if "****" in whatsapp or not whatsapp:
                whatsapp_clean = mobile
            else:
                # Take first number if multiple
                whatsapp_clean = whatsapp.split(",")[0].strip()

            # Parse rating
            try:
                rating = float(get("rating") or 0)
            except (ValueError, TypeError):
                rating = 0.0

            conn.execute("""
                INSERT INTO leads (
                    company_name, region, city, phone, mobile,
                    email, whatsapp, telegram, industry, rating
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                company_name,
                get("region"),
                get("city"),
                phone,
                mobile,
                get("email"),
                whatsapp_clean,
                get("telegram"),
                sheet_name,
                rating,
            ))
            total += 1

    conn.commit()
    wb.close()
    logger.info("Imported %d leads from %s", total, excel_path)
    return total


def get_leads_by_status(conn: sqlite3.Connection, status: str) -> list[dict]:
    """Get all leads with given status."""
    rows = conn.execute(
        "SELECT * FROM leads WHERE status = ? ORDER BY rating DESC", (status,)
    ).fetchall()
    return [dict(r) for r in rows]


def get_leads_by_industry(conn: sqlite3.Connection, industry: str) -> list[dict]:
    """Get all leads for an industry."""
    rows = conn.execute(
        "SELECT * FROM leads WHERE industry = ? ORDER BY rating DESC", (industry,)
    ).fetchall()
    return [dict(r) for r in rows]


def update_lead_status(conn: sqlite3.Connection, lead_id: int, status: str, notes: str = ""):
    """Update lead status and notes."""
    conn.execute(
        "UPDATE leads SET status = ?, notes = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (status, notes, lead_id),
    )
    conn.commit()


def get_stats(conn: sqlite3.Connection) -> dict:
    """Get funnel statistics."""
    rows = conn.execute(
        "SELECT status, COUNT(*) as cnt FROM leads GROUP BY status"
    ).fetchall()
    return {r["status"]: r["cnt"] for r in rows}


def get_industry_stats(conn: sqlite3.Connection) -> list[dict]:
    """Get stats per industry."""
    rows = conn.execute("""
        SELECT industry,
               COUNT(*) as total,
               SUM(CASE WHEN status = 'new' THEN 1 ELSE 0 END) as new_cnt,
               SUM(CASE WHEN status = 'called' THEN 1 ELSE 0 END) as called_cnt,
               SUM(CASE WHEN status = 'interested' THEN 1 ELSE 0 END) as interested_cnt,
               SUM(CASE WHEN status = 'sent_wa' THEN 1 ELSE 0 END) as sent_wa_cnt,
               SUM(CASE WHEN status = 'sent_email' THEN 1 ELSE 0 END) as sent_email_cnt,
               SUM(CASE WHEN status = 'done' THEN 1 ELSE 0 END) as done_cnt,
               SUM(CASE WHEN status = 'refused' THEN 1 ELSE 0 END) as refused_cnt
        FROM leads GROUP BY industry ORDER BY total DESC
    """).fetchall()
    return [dict(r) for r in rows]


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        path = os.environ.get("DEFAULT_EXCEL_PATH", "")
    if not path:
        print("Error: No Excel file specified. Pass as argument or set DEFAULT_EXCEL_PATH env var.")
        sys.exit(1)

    count = import_excel(path)
    print(f"Imported {count} leads")

    conn = init_db()
    stats = get_stats(conn)
    print(f"\nStats: {stats}")

    for row in get_industry_stats(conn):
        print(f"  {row['industry']}: {row['total']} total, {row['new_cnt']} new")
